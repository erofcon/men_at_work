from datetime import datetime, date
from tempfile import NamedTemporaryFile
from dateutil.relativedelta import relativedelta
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import permissions
from openpyxl import Workbook
from openpyxl.styles import Font
from task.models import Task
from openpyxl.writer.excel import save_virtual_workbook


class GetCountTask(APIView):
    """
    Generate statistics by months

    Returns data in a format:
            "labels": [
            "2022-02",
            "2022-03"
            ],
            "datasets": {
                "count_all_tasks": [
                    0,
                    2
                ],
                "count_executed_tasks": [
                    0,
                    0
                ],
                "count_expired_tasks": [
                    0,
                    2
                ]
            }

    requires the following parameters:
        start - start search date in a format: '2022-02-15 16:53'
        end - end search date in a format: '2022-02-15 16:53'

    """

    permission_classes = [permissions.IsAdminUser]

    def get(self, request):

        start = request.GET.get('start', False)
        end = request.GET.get('end', False)

        if start and end:
            start = start.rsplit('-', 1)[0]
            end = end.rsplit('-', 1)[0]

            start_date = datetime.strptime(start, "%Y-%m")
            end_date = datetime.strptime(end, "%Y-%m")

            labels = []

            count_all_tasks = []
            count_executed_tasks = []
            count_expired_tasks = []

            while start_date <= end_date:
                date_str = start_date.strftime("%Y-%m")
                labels.append(date_str)

                count_all_tasks.append(Task.objects.filter(createDateTime__gte=start_date).filter(
                    createDateTime__lte=start_date + relativedelta(months=1)).count())

                count_executed_tasks.append(Task.objects.filter(createDateTime__gte=start_date).
                                            filter(createDateTime__lte=start_date + relativedelta(months=1)).
                                            filter(is_done=True).count())

                count_expired_tasks.append(Task.objects.filter(createDateTime__gte=start_date).
                                           filter(createDateTime__lte=start_date + relativedelta(months=1)).
                                           filter(expired=True).filter(is_done=False).count())

                start_date += relativedelta(months=1)

            date_list = {
                'labels': labels,
                'datasets': {
                    'count_all_tasks': count_all_tasks,
                    'count_executed_tasks': count_executed_tasks,
                    'count_expired_tasks': count_expired_tasks,
                }
            }

            return Response(date_list)

        return Response({'parameters not specified'})


class CreateReport(APIView):
    """
    Create Report for admin user
    """

    permission_classes = [permissions.IsAdminUser]

    def get(self, request):

        start = request.GET.get('start', False)
        end = request.GET.get('end', False)

        if start and end:
            start = start.rsplit('-', 1)[0]
            end = end.rsplit('-', 1)[0]

            start_date = datetime.strptime(start, "%Y-%m")
            end_date = datetime.strptime(end, "%Y-%m")

            workbook = Workbook()
            excel_sheet = workbook.active

            excel_sheet['B4'] = f'Отчет с {start_date.strftime("%Y-%m")} по {end_date.strftime("%Y-%m")}'
            excel_sheet['B4'].font = Font(size=16)

            excel_sheet['C7'] = 'всего'
            excel_sheet['D7'] = 'исполнено'
            excel_sheet['E7'] = 'просрочено'

            i = 0
            while start_date <= end_date:
                date_str = start_date.strftime("%Y-%m")
                excel_sheet.cell(row=8+i, column=2).value = date_str

                excel_sheet.cell(row=8+i, column=3).value = (Task.objects.filter(createDateTime__gte=start_date).
                                                             filter(createDateTime__lte=start_date + relativedelta(months=1)).count())

                excel_sheet.cell(row=8+i, column=4).value = (Task.objects.filter(createDateTime__gte=start_date).
                                                             filter(createDateTime__lte=start_date + relativedelta(months=1)).
                                                             filter(is_done=True).count())

                excel_sheet.cell(row=8+i, column=5).value = (Task.objects.filter(createDateTime__gte=start_date).
                                                             filter(createDateTime__lte=start_date + relativedelta(months=1)).
                                                             filter(expired=True).filter(is_done=False).count())

                start_date += relativedelta(months=1)

                i += 1

            # temp_file = NamedTemporaryFile(delete = True)
            # workbook.save(temp_file)
            # stream = temp_file.read()
            # with NamedTemporaryFile(delete = True) as tmp:
            #     workbook.save(tmp.name)
            #     tmp.seek(0)
            #     stream = tmp.read()

            # with open('text.xls', 'wb') as tmp:
            #     workbook.save(tmp.name)
            #     # tmp.seek(0)
            #     stream = tmp.read()
            response = HttpResponse(save_virtual_workbook(workbook), content_type="application/ms-excel")

            # response = HttpResponse(content=stream, content_type='application/ms-excel', )
            response[
                'Content-Disposition'] = f'attachment; filename={datetime.now().strftime("%Y%m%d%H%M")}.xlsx'
            return response

        return Response({'parameters not specified'})
